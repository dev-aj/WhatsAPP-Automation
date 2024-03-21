#!/usr/bin/env python
# coding: utf-8

# installing pywhatkit for sending whatsapp messages
get_ipython().system('pip install pywhatkit')
# openpyxl for operting excel file
get_ipython().system('pip install openpyxl')


import pywhatkit as whatsapp
from openpyxl import Workbook, load_workbook
import time
import sys
import requests



path = '/Users/shad0w/Documents/Pyhton-Scripts/INVITATION.xlsx'
wb = load_workbook(path, read_only=True)
print(wb.sheetnames)
work = wb['Sheet1']

# Code to send images with caption text(Just checking if my code is working or not)
for row in result:
    contact_no = row
    image_path = "/Users/shad0w/Documents/Pyhton-Scripts/saraspdf.jpg"
    message = "Hello "
    wait_time = 10 # time to wait before sending the message (minute)
    close_tab = True # close the browser tab
    close_time = 20 # time to close the tab (minute)
    
    whatsapp.sendwhats_image(contact_no, image_path, message, wait_time, close_tab, close_time)
    time.sleep(30)


# Actual Loop to send messages
start = 571
end = 998
print("Start Time: ",time.localtime())
image_path = "/Users/shad0w/Documents/Pyhton-Scripts/photo.jpeg"
while start <= end:
    while not internet_connection():
        if internet_connection():
            print("The internet is connected")
            break
#         print("Not connected")
    cellName = 'B'+ str(start)
    c = work[cellName]
    phone = "+91"+str(c.value)
    comment = ''        
#     writeCell = 'R' + str(start)
    phone = "+91"+str(c.value)
    t = time.localtime()
    minute = t.tm_min+1
    hour = t.tm_hour
    
    comment = '''
        Hariom everyone,
A very warm greetings from Gyanoday Vidyalaya. 
We are elated to announce that like every year, this year also Gyanoday’s students have outperformed everyone in the Sainik💂 School🏫 admission examination. Have an overview of our result 🏆 by yourself.
•100% (45/45) students qualified
•AIR-23, Class 9
•Under 100 ranks: AIR-23, AIR-50 AIR-63, AIR-93
To know more, kindly contact on below 👇 phone no📱 9835390369 / 9934901178\n
हरिओम सबको,
ज्ञानोदय विद्यालय की ओर से हार्दिक शुभकामनाएं।
हमें यह घोषणा करते हुए खुशी हो रही है कि हर साल की तरह इस साल भी ज्ञानोदय के छात्रों ने सैनिक स्कूल प्रवेश परीक्षा में सभी से बेहतर प्रदर्शन किया है। हमारे परिणाम का अवलोकन 🏆 स्वयं करें।
•100% (45/45) छात्र उत्तीर्ण हुए
•All India Rank-23, कक्षा 9
•100 से कम रैंक: AIR-23, AIR-50 AIR-63, AIR-93
अधिक जानने के लिए कृपया नीचे दिए गए 👇 फोन नंबर📱 9835390369 / 9934901178 पर संपर्क करें।'''
#     print(comment)
   # break;
    if(minute >= 60):
        minute = minute % 60 + 2
        hour = hour + 1
    if(c.value != None and comment != 'Operation Failed'):
        phone = "+91"+str(c.value)
        whatsapp.sendwhats_image(phone, image_path, comment,15, True, 5)
#         whatsapp.sendwhats_image(phone, "C:/Users/AJ/Downloads/QRTrust.jpeg", "QR-CODE FOR PAYMENT", 20,True,3)
#         c4 = work[writeCell]
#         c4.value = comment
    elif(c.value == None or comment == 'Operation Failed'):
        comment = '''Operation Failed OR Mobile number not found for.'''
        phone = "+919955599298"
        whatsapp.sendwhatmsg(phone, comment, hour, minute, 10, True, 4)

    print("Adm : ",start)
    start += 1

    
print("End Time: ",time.localtime())
