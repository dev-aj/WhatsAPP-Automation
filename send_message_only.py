#!/usr/bin/env python
# coding: utf-8

# installing pywhatkit for sending whatsapp messages
get_ipython().system('pip install pywhatkit')
# openpyxl for operting excel file
get_ipython().system('pip install openpyxl')


import pywhatkit as whatsapp
from openpyxl import Workbook, load_workbook
import time
import sys
import requests

# Importing excel file named Call.xlsx
path = '/Users/shad0w/Documents/Pyhton-Scripts/Calls.xlsx'
wb = load_workbook(path, read_only=True)
print(wb.sheetnames)
work = wb['Sheet4']


# Just to check if the file has been imported or not
c = work['B2'].value
phone = "+91"+str(c)
print(phone)

# To check if system is connected to internet (as it is required to send whatssapp messages)
def internet_connection():
    try:
        response = requests.get("https://web.whatsapp.com/", timeout=5)
        return True
    except requests.ConnectionError:
        return False
if internet_connection():
    print("The internet is connected1")
else:
    print("Not connected1")

# Here the loop starts to sending messages
start = 3
end = 318
print("Start Time: ",time.localtime())
while start <= end:
    while not internet_connection():
        if internet_connection():
            print("The internet is connected")
            break
    cellName = 'E'+ str(start)
    admcell = 'C' + str(start)
    adm = work[admcell]
    c = work[cellName]
    phone = "+91"+str(c.value)
    comment = ''
    old_fee = 0
    try:
        dues = work['P'+ str(start)].value
        ward = work['B' + str(start)].value
        month = 'March'
  
    except:
        comment = "Operation Failed"
    #print(comment)
        
    writeCell = 'O' + str(start)
    phone = "+91"+str(c.value)
    t = time.localtime()
    minute = t.tm_min + 1
    hour = t.tm_hour

    if(dues <= 0):
        comment = '''
        हरी ओम ! यह ज्ञानोदय विद्यालय की ओर से एक संदेश है।

हमें आगामी होली की छुट्टियों की घोषणा करते हुए खुशी हो रही है! छुट्टियां 16 मार्च से शुरू होकर 27 मार्च तक चलेंगी।

अपने कैलेंडर में चिन्हित करें!
1. सभी छात्रों को 28 से 29 मार्च तक विद्यालय कैंपस में पहुंचना अनिवार्य है।
2. अभिभावक-शिक्षक बैठकें (PTM) 28 मार्च और 29 मार्च को सुबह 9:00 बजे से दोपहर 2:00 बजे तक आयोजित की जाएंगी। 

आप और आपके परिवार को होली की शुभकामनाएं! ✨

हार्दिक शुभकामनाएं,

ज्ञानोदय विद्यालय टीम'''
    else:
        comment = '''
            हरी ओम ! यह ज्ञानोदय विद्यालय की ओर से एक संदेश है।

हमें आगामी होली की छुट्टियों की घोषणा करते हुए खुशी हो रही है! छुट्टियां 16 मार्च से शुरू होकर 27 मार्च तक चलेंगी।

*15 मार्च से पहले अपना नो ड्यू सर्टिफिकेट लेना सुनिश्चित करें और अंत समय में होने वाली परेशानियों से बचें।*

अपने कैलेंडर में चिन्हित करें!
1. सभी छात्रों को 28 से 29 मार्च तक विद्यालय कैंपस में पहुंचना अनिवार्य है।
2. अभिभावक-शिक्षक बैठकें (PTM) 28 मार्च और 29 मार्च को सुबह 9:00 बजे से दोपहर 2:00 बजे तक आयोजित की जाएंगी। 

आप और आपके परिवार को होली की शुभकामनाएं! ✨

हार्दिक शुभकामनाएं,

ज्ञानोदय विद्यालय टीम'''
    if(minute >= 60):
        minute = minute % 60 + 2
        hour = hour + 1
    if(c.value != None and comment != 'Operation Failed'):
        phone = "+91"+str(c.value)
        whatsapp.sendwhatmsg(phone, comment, hour, minute,15, True, 10)
#         c4 = work[writeCell]
#         c4.value = comment
    elif(c.value == None or comment == 'Operation Failed'):
        comment = '''Operation Failed OR Mobile number not found for {ward}'''.format(ward = ward)
        phone = "+919955599298"
        whatsapp.sendwhatmsg(phone, comment, hour, minute, 10, True, 4)
#         c4 = work[writeCell]
#         c4.value = comment
        
#     wb.save(path)
    print('Sr.No :',start)
    start += 1

    
print("End Time: ",time.localtime())